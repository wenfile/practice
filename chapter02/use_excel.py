from datetime import datetime

import MySQLdb
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors


class ExcelUtils(object):

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_two = self.wb.create_sheet('我的表单')
        self.ws_three = self.wb.create_sheet()
        self.ws_two.title = '你的表单'
        self.ws_two.sheet_properties.tabColor = '00ff00'
        self.ws_four = self.wb.create_sheet('新的表单')

    def do_sth(self):
        # 插入数据
        self.ws['A1'] = '666'
        self.ws['A2'] = 'hello'
        self.ws['A3'] = datetime.now()

        # 插入图片
        # img = Image('../static/temp.png')
        # self.ws.add_image(img, 'B1')

        # 设置字体样式
        font = Font(sz=18, color=colors.RED, b=1)
        self.ws['A1'].font = font

        # 合并单元格
        self.ws.merge_cells('A4:E5')
        self.ws.unmerge_cells('A4:E5')

        i = 1
        for row in self.ws_two['A1:D4']:
            for cell in row:
                cell.value = i
                i = i + 1
            i = 1

        # 使用excel公式
        self.ws_two['e1'] = '=sum(A1:D1)'

        self.wb.save('../static/test.xlsx')

    def read_excel(self):
        """
         读取excel
         """
        wb = load_workbook('../static/grade.xlsx')
        sheets = wb.get_sheet_names()
        print(sheets)
        wb = wb.active
        for (i, row) in enumerate(wb.rows):
            if i < 2:
                continue
            year = wb['A{0}'.format(i + 1)].value
            max_grad = wb['B{0}'.format(i + 1)].value
            avg = wb['C{0}'.format(i + 1)].value
            # print(year, max_grad, avg)
            conn = self.get_conn()
            cursor = conn.cursor()
            sql = 'INSERT INTO `score`(`year`,`max`,`avg`) VALUES ({year}, {max_grad}, {avg})'.format(year=year, max_grad=max_grad, avg=avg)
            print(sql)
            cursor.execute(sql)
            conn.autocommit(True)
            # print(conn)

    def get_conn(self):
        conn = MySQLdb.connect(
            db='user_grade',
            host='localhost',
            user='root',
            password='',
            charset='utf8'
        )
        return conn

    def export_conn(self):
        conn = self.get_conn()
        cursor = conn.cursor()
        sql = 'SELECT `year`, `max`, `avg` FROM `score`'
        cursor.execute(sql)
        rows = cursor.fetchall()
        # print(rows)
        wb = load_workbook('../static/export.xlsx')
        ws = wb.active
        for (i, row) in enumerate(rows):
            print(row)
            # ws['A1'].value = 'year'
            # ws['B1'].value = 'max'
            # ws['C1'].value = 'avg'
            (ws['A{0}'.format(i+2)].value,
                ws['B{0}'.format(i+2)].value,
                ws['C{0}'.format(i+2)].value) = row
        wb.save('../static/export.xlsx')


if __name__ == '__main__':
    client = ExcelUtils()
    # client.do_sth()
    # client.read_excel()
    client.export_conn()